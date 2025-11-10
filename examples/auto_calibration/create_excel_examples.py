#!/usr/bin/env python3
"""
Create Excel protocol examples for automatic calibration.

These examples demonstrate the new automatic calibration system without
manual CALIBRATION_FACTOR in the Excel file.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

def create_simple_blink_excel():
    """Create simple_blink_example.xlsx with automatic calibration."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # ========================================================================
    # PROTOCOL SHEET
    # ========================================================================
    ws_protocol = wb.create_sheet("protocol")
    
    # Header row
    headers = ['time_sec', 'CH1_status', 'CH1_time_sec', 'CH1_period', 'CH1_pulse_width',
               'CH2_status', 'CH2_time_sec', 'CH2_period', 'CH2_pulse_width']
    ws_protocol.append(headers)
    
    # Style header
    for cell in ws_protocol[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Channel 1: Simple blink (ON 1s, OFF 1s) × 30 repeats = 60 rows
    time = 0
    for i in range(30):
        # ON state
        ws_protocol.append([time, 1, 1, 0, 0, 0, 0, 0, 0])
        time += 1
        # OFF state
        ws_protocol.append([time, 0, 1, 0, 0, 0, 0, 0, 0])
        time += 1
    
    # Channel 2: Pulsed pattern (5Hz pulse for 2s, OFF for 2s) × 15 repeats = 30 rows
    time = 0
    for i in range(15):
        # Pulse ON (5Hz = 200ms period, 20ms pulse width)
        ws_protocol.append([time, 0, 0, 0, 0, 1, 2, 200, 20])
        time += 2
        # OFF state
        ws_protocol.append([time, 0, 0, 0, 0, 0, 2, 0, 0])
        time += 2
    
    # ========================================================================
    # START_TIME SHEET
    # ========================================================================
    ws_start = wb.create_sheet("start_time")
    
    # Column-based format (modern)
    ws_start.append(['Channels', 'start_time', 'wait_status'])
    ws_start.append(['CH1', 0, 0])
    ws_start.append(['CH2', 5, 0])
    
    # Style header
    for cell in ws_start[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # ========================================================================
    # CALIBRATION SHEET - INTENTIONALLY OMITTED FOR AUTO-CALIBRATION
    # ========================================================================
    # Note: We do NOT create a calibration sheet.
    # This signals to the system to use automatic calibration.
    
    # ========================================================================
    # INFO SHEET (Optional - explains automatic calibration)
    # ========================================================================
    ws_info = wb.create_sheet("info")
    ws_info.column_dimensions['A'].width = 80
    
    info_text = [
        ["AUTOMATIC CALIBRATION EXAMPLE"],
        [""],
        ["This Excel protocol uses the NEW automatic calibration system."],
        [""],
        ["KEY DIFFERENCES FROM PRESET CALIBRATION:"],
        ["  ✗ NO 'calibration' sheet included"],
        ["  ✓ System automatically identifies your Arduino board"],
        ["  ✓ Calibration factor retrieved from database"],
        ["  ✓ First-time users prompted to calibrate once"],
        ["  ✓ Subsequent runs are automatic - no manual tracking"],
        [""],
        ["FIRST RUN WORKFLOW:"],
        ["  1. System identifies Arduino by serial number/VID:PID"],
        ["  2. Checks calibration database"],
        ["  3. If not calibrated, prompts: 'Calibrate now? (Y/n)'"],
        ["  4. Calibration saved automatically to database"],
        ["  5. Protocol execution begins with calibrated timing"],
        [""],
        ["SUBSEQUENT RUNS:"],
        ["  - Calibration loaded automatically from database"],
        ["  - No prompts or manual input needed"],
        ["  - Works even if Arduino plugged into different USB port"],
        [""],
        ["WHY NO CALIBRATION SHEET?"],
        ["  The absence of a 'calibration' sheet signals to the parser that"],
        ["  this protocol uses automatic calibration. The system will:"],
        ["    1. Detect the connected Arduino board"],
        ["    2. Look up its unique ID in the database"],
        ["    3. Apply the stored calibration factor"],
        [""],
        ["BENEFITS:"],
        ["  • No manual calibration factor management"],
        ["  • Seamless multi-board support"],
        ["  • Eliminates tracking errors"],
        ["  • Simpler Excel file structure"],
        [""],
        ["MANAGEMENT COMMANDS:"],
        ["  View calibrations:   python utils/manage_calibrations.py list"],
        ["  Test board ID:       python test_board_info.py"],
        ["  Force recalibrate:   python protocol_parser.py 2 <port> <file> --calibrate"],
        [""],
        ["DOCUMENTATION:"],
        ["  Complete guide: docs/AUTO_CALIBRATION_DATABASE.md"],
        ["  Compatibility:  docs/BACKWARD_COMPATIBILITY.md"],
        ["  TXT examples:   examples/auto_calibration/simple_blink_example.txt"],
    ]
    
    for row in info_text:
        ws_info.append(row)
    
    # Style info sheet
    ws_info['A1'].font = Font(bold=True, size=14)
    ws_info['A5'].font = Font(bold=True)
    ws_info['A12'].font = Font(bold=True)
    ws_info['A18'].font = Font(bold=True)
    ws_info['A20'].font = Font(bold=True)
    ws_info['A28'].font = Font(bold=True)
    ws_info['A32'].font = Font(bold=True)
    ws_info['A37'].font = Font(bold=True)
    
    # Save
    output_path = os.path.join(os.path.dirname(__file__), 'simple_blink_example.xlsx')
    wb.save(output_path)
    print(f"✓ Created: {output_path}")


def create_pulse_protocol_excel():
    """Create pulse_protocol.xlsx with automatic calibration."""
    wb = Workbook()
    wb.remove(wb.active)
    
    # ========================================================================
    # PROTOCOL SHEET
    # ========================================================================
    ws_protocol = wb.create_sheet("protocol")
    
    # Header - 4 channels with pulse parameters
    headers = ['time_sec']
    for ch in range(1, 5):
        headers.extend([f'CH{ch}_status', f'CH{ch}_time_sec', f'CH{ch}_period', f'CH{ch}_pulse_width'])
    ws_protocol.append(headers)
    
    # Style header
    for cell in ws_protocol[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Create pulsed patterns for each channel
    # CH1: 1Hz pulse (T1000pw100) for 5s, OFF for 5s, repeat 6 times
    time = 0
    for i in range(6):
        ws_protocol.append([time, 1, 5, 1000, 100, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
        time += 5
        ws_protocol.append([time, 0, 5, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
        time += 5
    
    # CH2: 2Hz pulse (T500pw50) for 4s, OFF for 4s, repeat 8 times
    time = 0
    for i in range(8):
        ws_protocol.append([time, 0, 0, 0, 0, 1, 4, 500, 50, 0, 0, 0, 0, 0, 0, 0, 0])
        time += 4
        ws_protocol.append([time, 0, 0, 0, 0, 0, 4, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
        time += 4
    
    # CH3: 5Hz pulse (T200pw20) for 3s, OFF for 3s, repeat 10 times
    time = 0
    for i in range(10):
        ws_protocol.append([time, 0, 0, 0, 0, 0, 0, 0, 0, 1, 3, 200, 20, 0, 0, 0, 0])
        time += 3
        ws_protocol.append([time, 0, 0, 0, 0, 0, 0, 0, 0, 0, 3, 0, 0, 0, 0, 0, 0])
        time += 3
    
    # CH4: 10Hz pulse (T100pw10) for 2s, OFF for 2s, repeat 15 times
    time = 0
    for i in range(15):
        ws_protocol.append([time, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 2, 100, 10])
        time += 2
        ws_protocol.append([time, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2, 0, 0])
        time += 2
    
    # ========================================================================
    # START_TIME SHEET
    # ========================================================================
    ws_start = wb.create_sheet("start_time")
    ws_start.append(['Channels', 'start_time', 'wait_status'])
    ws_start.append(['CH1', 0, 0])
    ws_start.append(['CH2', 5, 0])
    ws_start.append(['CH3', 10, 0])
    ws_start.append(['CH4', 15, 0])
    
    for cell in ws_start[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # ========================================================================
    # INFO SHEET
    # ========================================================================
    ws_info = wb.create_sheet("info")
    ws_info.column_dimensions['A'].width = 80
    
    info_text = [
        ["PULSE PROTOCOL - AUTOMATIC CALIBRATION"],
        [""],
        ["This example demonstrates pulsed patterns with automatic calibration."],
        [""],
        ["PULSE PATTERNS:"],
        ["  CH1: 1Hz pulse (1000ms period, 100ms width) - slow blink"],
        ["  CH2: 2Hz pulse (500ms period, 50ms width) - medium blink"],
        ["  CH3: 5Hz pulse (200ms period, 20ms width) - fast blink"],
        ["  CH4: 10Hz pulse (100ms period, 10ms width) - rapid blink"],
        [""],
        ["AUTOMATIC CALIBRATION:"],
        ["  No 'calibration' sheet = automatic calibration enabled"],
        ["  System identifies Arduino and applies stored calibration"],
        ["  Pulse timing accuracy ensured through calibrated factors"],
        [""],
        ["PULSE TIMING IMPORTANCE:"],
        ["  Accurate pulse timing requires precise calibration."],
        ["  The automatic system ensures each Arduino's unique clock"],
        ["  characteristics are properly compensated."],
        [""],
        ["For more information: docs/AUTO_CALIBRATION_DATABASE.md"],
    ]
    
    for row in info_text:
        ws_info.append(row)
    
    ws_info['A1'].font = Font(bold=True, size=14)
    ws_info['A5'].font = Font(bold=True)
    ws_info['A11'].font = Font(bold=True)
    ws_info['A15'].font = Font(bold=True)
    
    output_path = os.path.join(os.path.dirname(__file__), 'pulse_protocol.xlsx')
    wb.save(output_path)
    print(f"✓ Created: {output_path}")


def create_multi_channel_excel():
    """Create multi_channel_pattern.xlsx with automatic calibration."""
    wb = Workbook()
    wb.remove(wb.active)
    
    # ========================================================================
    # PROTOCOL SHEET
    # ========================================================================
    ws_protocol = wb.create_sheet("protocol")
    
    # Header - 4 channels
    headers = ['time_sec']
    for ch in range(1, 5):
        headers.extend([f'CH{ch}_status', f'CH{ch}_time_sec', f'CH{ch}_period', f'CH{ch}_pulse_width'])
    ws_protocol.append(headers)
    
    for cell in ws_protocol[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # CH1: Four-phase cycle (ON 1s, OFF 3s, ON 2s, OFF 2s) × 10 repeats
    time = 0
    for i in range(10):
        ws_protocol.append([time, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
        time += 1
        ws_protocol.append([time, 0, 3, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
        time += 3
        ws_protocol.append([time, 1, 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
        time += 2
        ws_protocol.append([time, 0, 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
        time += 2
    
    # CH2: Pulse intensity ramp (low→medium→high→off) × 5 repeats
    time = 0
    for i in range(5):
        # Low intensity (10% duty cycle)
        ws_protocol.append([time, 0, 0, 0, 0, 1, 3, 1000, 100, 0, 0, 0, 0, 0, 0, 0, 0])
        time += 3
        # Medium (30%)
        ws_protocol.append([time, 0, 0, 0, 0, 1, 3, 1000, 300, 0, 0, 0, 0, 0, 0, 0, 0])
        time += 3
        # High (50%)
        ws_protocol.append([time, 0, 0, 0, 0, 1, 3, 1000, 500, 0, 0, 0, 0, 0, 0, 0, 0])
        time += 3
        # Off
        ws_protocol.append([time, 0, 0, 0, 0, 0, 3, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
        time += 3
    
    # CH3: Heartbeat (pulse-pause-pulse-rest) × 20 repeats
    time = 0
    for i in range(20):
        ws_protocol.append([time, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0.2, 100, 50, 0, 0, 0, 0])
        time += 0.2
        ws_protocol.append([time, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0.2, 0, 0, 0, 0, 0, 0])
        time += 0.2
        ws_protocol.append([time, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0.2, 100, 50, 0, 0, 0, 0])
        time += 0.2
        ws_protocol.append([time, 0, 0, 0, 0, 0, 0, 0, 0, 0, 2.4, 0, 0, 0, 0, 0, 0])
        time += 2.4
    
    # CH4: Variable frequency (1Hz→2Hz→5Hz→10Hz) × 3 repeats
    time = 0
    for i in range(3):
        ws_protocol.append([time, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 4, 1000, 100])
        time += 4
        ws_protocol.append([time, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 4, 500, 50])
        time += 4
        ws_protocol.append([time, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 4, 200, 20])
        time += 4
        ws_protocol.append([time, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 4, 100, 10])
        time += 4
    
    # ========================================================================
    # START_TIME SHEET
    # ========================================================================
    ws_start = wb.create_sheet("start_time")
    ws_start.append(['Channels', 'start_time', 'wait_status'])
    ws_start.append(['CH1', 0, 0])
    ws_start.append(['CH2', 5, 0])
    ws_start.append(['CH3', 10, 0])
    ws_start.append(['CH4', 15, 1])  # CH4 waits in ON state
    
    for cell in ws_start[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # ========================================================================
    # INFO SHEET
    # ========================================================================
    ws_info = wb.create_sheet("info")
    ws_info.column_dimensions['A'].width = 80
    
    info_text = [
        ["MULTI-CHANNEL PATTERN - AUTOMATIC CALIBRATION"],
        [""],
        ["Complex multi-channel coordination with automatic calibration."],
        [""],
        ["CHANNEL PATTERNS:"],
        ["  CH1: Four-phase cycle (short ON, long OFF, medium ON, medium OFF)"],
        ["  CH2: Pulse intensity ramp (10% → 30% → 50% → OFF duty cycles)"],
        ["  CH3: Heartbeat pattern (double-pulse with rest period)"],
        ["  CH4: Variable frequency sweep (1Hz → 2Hz → 5Hz → 10Hz)"],
        [""],
        ["AUTOMATIC CALIBRATION:"],
        ["  This Excel file intentionally omits the 'calibration' sheet."],
        ["  The system will automatically apply board-specific calibration."],
        [""],
        ["WHY BOARD-SPECIFIC CALIBRATION MATTERS:"],
        ["  Each Arduino board has unique clock characteristics due to:"],
        ["    • Crystal oscillator tolerances (±50-100 ppm typical)"],
        ["    • Temperature-dependent frequency drift"],
        ["    • Manufacturing variations in crystal load capacitance"],
        ["    • Component aging effects"],
        [""],
        ["  Even two 'identical' Arduino boards may have timing differences"],
        ["  of 0.1-1.0%, which accumulates over long protocols."],
        [""],
        ["  Automatic calibration measures YOUR specific board's timing"],
        ["  and stores a correction factor unique to that board."],
        [""],
        ["For detailed information: docs/AUTO_CALIBRATION_DATABASE.md"],
    ]
    
    for row in info_text:
        ws_info.append(row)
    
    ws_info['A1'].font = Font(bold=True, size=14)
    ws_info['A5'].font = Font(bold=True)
    ws_info['A11'].font = Font(bold=True)
    ws_info['A15'].font = Font(bold=True)
    
    output_path = os.path.join(os.path.dirname(__file__), 'multi_channel_pattern.xlsx')
    wb.save(output_path)
    print(f"✓ Created: {output_path}")


if __name__ == '__main__':
    print("Creating Excel examples for automatic calibration...")
    print()
    
    create_simple_blink_excel()
    create_pulse_protocol_excel()
    create_multi_channel_excel()
    
    print()
    print("✓ All Excel examples created successfully!")
    print()
    print("These Excel files demonstrate automatic calibration:")
    print("  • No 'calibration' sheet included")
    print("  • System automatically identifies Arduino board")
    print("  • Calibration factor retrieved from database")
    print("  • One-time setup per Arduino board")
    print()
    print("To use these examples:")
    print("  python protocol_parser.py 2 <port> auto_calibration/simple_blink_example.xlsx")
    print("  python protocol_parser.py 4 <port> auto_calibration/pulse_protocol.xlsx")
    print("  python protocol_parser.py 4 <port> auto_calibration/multi_channel_pattern.xlsx")
